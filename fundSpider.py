#
#   @Author Hypocritical Fish
#   @Create 2022/10/15 21:51
#   @Description    爬取天天基金投顾产品的数据


import requests
import time
import openpyxl
import os

import sys

from loggers import get_logger

from urllib3.exceptions import InsecureRequestWarning

from utils import get_headers, get_timestamp, del_file

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

log = get_logger(True)


# 获取投顾产品id列表
def get_partner_id_list():
    url = 'https://universalapi.1234567.com.cn/hxfundtrade/auxiliary/kycPartnerList'
    host = 'universalapi.1234567.com.cn'
    referer = 'https://mpservice.com/b0381a3b634440379d330b69f09d3f8e/release/pages/selectStrategy/index'
    headers = get_headers(host, referer, '')
    params = {
        'headers': '{content-type=application/json}'
    }
    response = requests.get(url=url, params=params, headers=headers, verify=False)
    response.encoding = 'utf-8'
    partner_id_list = response.json().get('data')
    log.info("投顾产品总数：" + str(len(partner_id_list)))
    return partner_id_list


# 根据partnerId获取基本信息
def get_base_info(partner_id):
    url = 'https://uni-fundts.1234567.com.cn/combine/investAdviserInfo/getAgencyConfigInfo'
    host = 'uni-fundts.1234567.com.cn'
    referer = 'https://mpservice.com/funda91a99886abf7e/release/pages/question/index?partnerId=' + partner_id
    headers = get_headers(host, referer, '')
    params = {
        'partnerId': partner_id,
    }
    response = requests.get(url=url, params=params, headers=headers, verify=False)
    response.encoding = 'utf-8'
    data = response.json().get('Data')
    item = {
        'partnerId': data['partnerId'],
        'companyId': data['companyId'],
        'partnerName': data['appletName'],
        'inRateInfo': data['inRateInfo']
    }
    return item


# 获取指定id的投顾产品的策略id集合
def get_strategy_id_set(partner_id):
    url = 'https://universalapi.1234567.com.cn/hxfundtrade/auxiliary/kycQuestionV2'
    host = 'universalapi.1234567.com.cn'
    referer = 'https://mpservice.com/funda91a99886abf7e/release/pages/question/index?partnerId=' + partner_id
    headers = get_headers(host, referer, '')
    params = {
        'partner': partner_id,
    }
    response = requests.get(url=url, params=params, headers=headers, verify=False)
    response.encoding = 'utf-8'
    content = response.json().get('data')
    log.info('开始爬取' + content['investConsultName'] + '的投顾策略...')
    strategy_id_set = set()
    for item in content['ruleList']:
        strategy_id_list = item['strategyId'].replace('*', '').replace('&', ',').split(',')
        for strategy_id in strategy_id_list:
            if strategy_id != '0':
                strategy_id_set.add(strategy_id)
    log.info('爬取' + content['investConsultName'] + '旗下共计' + str(len(strategy_id_set)) + '条投顾策略...')
    return strategy_id_set


# 获取策略全部信息
def get_strategy_info(item, strategy_id):
    url = 'https://h5.1234567.com.cn/AggregationStaticService/chooseCustomRouter/getStrategyDetailAggr'
    host = 'h5.1234567.com.cn'
    referer = 'https://mpservice.com/funda91a99886abf7e/release/pages/strategyDetail/index?partnerId=' + item[
        'partnerId'] + '&id=' + strategy_id
    headers = get_headers(host, referer, 'application/x-www-form-urlencoded')
    data = {
        'partner': item['partnerId'],
        'MobileKey': 'e952ac58c6c75cb128a59455b39a0b24||iemi_tluafed_me',
        'UserId': '95c68e3ae0884f828ecedea62eb3e24c',
        'PassportId': '2884496638254532',
        'UToken': 'T2C6uy5YFLoZPtBR-oOL1orUhSKBUIszzKk6vaI6ku7voy6Zf6IfNupjnditqV_RzgSj0Zz4XY0JUdNVbTd5J7EvcDAHmcb2JR9JQHjyIugNb_jj7FOjDbxEe89ErH28dUsxzQFtg92npSVcETB6DGDq9MKWkjcnbjvXs01zAAg.12',
        'tag': '0',
        'displayId': strategy_id,
        'JJGSID': item['companyId'],
        'tgCode': strategy_id,
        'CToken': 'dvPbcblCxG2nGQGU1msJDSQOcDdWmwG9wOJchTPrqhvoOqkWiNShgsapqwdQNVvpZJSMq3KG4hRe94ruyJ2SmDEbJMna8VOdX763fFasdKw1omuMb6DmnKDK1VdI6RFluxzsz-_8rBlb9DSyqj-XrUHRPlSu8eP-Aj60BOuGzAfuO4U5Jmluj7Z0ZRr1J6w_ZtpsGN9skZ-zpZpljuga5mDkSbcBIB3qWzUdB0YPumA.12',
        'TGCODE': strategy_id,
    }
    response = requests.post(url=url, data=data, headers=headers, verify=False)
    response.encoding = 'utf-8'
    content = response.json()['data']
    return content


# 解析基本策略数据
def parse_strategy_basic(strategy_id, item, partner_ws, content):
    # 基本数据
    base_data = content['baseData']['data']
    # 基本数据为空则不解析
    if base_data is None or len(base_data) == 0:
        global skip_count
        skip_count = skip_count + 1
        log.warning(item[
                        'partnerName'] + "公司旗下id为" + strategy_id + "的投顾策略基本信息缺失，无法解析!即将解析下一条投顾策略...")
        return False

    # 策略基本信息
    strategy_info_aggr = content['strategyInfoAggr']['data']

    # 目标收益信息
    target_profit_info = content['targetProfitInfo']['data']
    # 工作日列表
    work_day_list = content['workdayList']['data']['workDayLst']

    # 投顾策略名称
    strategy_name = base_data['strategyName']
    item['strategyName'] = strategy_name

    # 推荐持有
    recommend_hold_time = base_data.get('recommendHoldTime', '')
    item['recommendHoldTime'] = recommend_hold_time

    key_title1 = base_data.get('keyTitle1', '')
    hold_limit1 = base_data.get('holdLimit1', '')
    item['keyTitle1'] = key_title1
    item['holdLimit1'] = hold_limit1

    # 成立天数
    continued_data = strategy_info_aggr[0].get('continuedData', '')
    item['continuedData'] = continued_data

    # 成立来业绩表现
    total_performance = content['strategyProfitChart_ln'].get('data', {})
    length = len(total_performance) - 1
    if length > 0:
        item['totalPerformance'] = total_performance[length]['SE'] + "%"
        item['totalBasic'] = total_performance[length]['BENCH_SE'] + "%"

    # 昨日涨幅
    yesterday_performance = target_profit_info[0].get('SYL_D', '')
    if yesterday_performance is None:
        item['yesterdayPerformance'] = ''
    else:
        item['yesterdayPerformance'] = yesterday_performance + "%"

    # 月业绩表现
    month_performance = content['strategyProfitChart_y'].get('data', {})
    length = len(month_performance) - 1
    if length > 0:
        item['monthPerformance'] = month_performance[length]['SE'] + "%"
        item['monthBasic'] = month_performance[length]['BENCH_SE'] + "%"

    # 近三月业绩表现
    three_month_performance = content['strategyProfitChart_3y']['data']
    length = len(three_month_performance) - 1
    if length > 0:
        item['threeMonthPerformance'] = three_month_performance[length]['SE'] + "%"
        item['threeMonthBasic'] = three_month_performance[length]['BENCH_SE'] + "%"

    # 近半年业绩表现
    six_month_performance = content['strategyProfitChart_6y']['data']
    length = len(six_month_performance) - 1
    if length > 0:
        item['sixMonthPerformance'] = six_month_performance[length]['SE'] + "%"
        item['sixMonthBasic'] = six_month_performance[length]['BENCH_SE'] + "%"

    # 近一年业绩表现
    year_performance = content['strategyProfitChart_n']['data']
    length = len(year_performance) - 1
    if length > 0:
        item['yearPerformance'] = year_performance[length]['SE'] + "%"
        item['yearBasic'] = year_performance[length]['BENCH_SE'] + "%"

    # 服务费率
    fee_rate_info = strategy_info_aggr[0].get('STRATEGY_RATE')
    item['serviceRate'] = str(float(fee_rate_info[fee_rate_info.index("0."):]) * 100) + "%"

    # 最新调仓动态，历史调仓未解析
    adjust_warehouse_info = content['getAdjustWarehouse_0'].get('data', {})
    adjust_date = adjust_warehouse_info['latestAdjust']['dateStr']
    reason = adjust_warehouse_info['latestAdjust']['reason']
    item['lastAdjustDate'] = adjust_date
    item['lastAdjustReason'] = reason

    partner_ws.append((
        item.get('partnerName', ''),
        item.get('strategyName', ''),
        item.get('totalPerformance', ''),
        item.get('totalBasic', ''),
        item.get('yesterdayPerformance', ''),
        item.get('continuedData', ''),
        item.get('recommendHoldTime', ''),
        item.get('holdLimit1', ''),
        item.get('monthPerformance', ''),
        item.get('monthBasic', ''),
        item.get('threeMonthPerformance', ''),
        item.get('threeMonthBasic', ''),
        item.get('sixMonthPerformance', ''),
        item.get('sixMonthBasic', ''),
        item.get('yearPerformance', ''),
        item.get('yearBasic', ''),
        item.get('inRateInfo', ''),
        item.get('serviceRate', ''),
        item.get('lastAdjustDate', ''),
        item.get('lastAdjustReason', ''),
    ))
    log.info('\t' + strategy_name + '基本信息爬取成功')
    return True


# 解析详细策略数据
def parse_strategy_detail(strategy_ws, content):
    # 持仓信息
    hold_warehouse_info = content['getHoldWarehouseInfo'].get('data', {})
    # 持仓基金类型列表
    fund_type_map = ("QDII", "股票型", "货币型", "混合型", "", "", "债券型", "", "指数型")
    hold_type_list = hold_warehouse_info.get('holdTypeList', [])
    detail = ''

    ratios_info = content['strategyModel']['data'][0]['POSITION'].split(',')
    for i, value in enumerate(ratios_info):
        if i % 2 == 1:
            pre = ratios_info[i - 1]
            j = pre.rfind('"')
            k = pre[j - 1:j]
            if k == 'a':
                k = '0'
            fund_type = int(k)
            left = value.index(':') + 1
            right = value.index('}')
            ratio = round(float(value[left:right]) * 100, 2)
            detail += fund_type_map[fund_type] + "基金总占比：" + str(ratio) + "%\t\n"
    strategy_ws.append(('【持仓分布】',))
    strategy_ws.append((detail,))
    if hold_type_list is None or len(hold_type_list) == 0:
        strategy_ws.append(())
        strategy_ws.append(('无持仓分布详情',))
    else:
        strategy_ws.append(())
        strategy_ws.append(('【持仓详情】',))
        strategy_ws.append(('基金代码', '基金名称', '占比', '类型', '日期', '净值', '涨幅'))
        for holdType in hold_type_list:
            for fund in holdType['fundsList']:
                # 计算类型
                t = fund['type']
                if t == 'a':
                    t = '0'
                fund_type = fund_type_map[int(t)]

                increase = fund.get('increase', '')
                if increase is None:
                    increase = ''
                else:
                    increase = increase + '%'

                strategy_ws.append((
                    fund.get("fundCode", ''),
                    fund.get("fundName", ''),
                    fund['ratio'] + "%",
                    fund_type,
                    fund.get("date", ''),
                    fund.get("netAssetValue", ''),
                    increase
                ))

    # 调仓详情
    adjust_warehouse_info = content['getAdjustWarehouse_0'].get('data', {})
    adjust_list = adjust_warehouse_info['latestAdjust']['adjustList']
    detail = ''
    for adjustFund in adjust_list:
        t = adjustFund['type']
        if t == 'a':
            t = '0'
        detail += fund_type_map[int(t)] + "基金总占比：" + adjustFund['preTotalRatio'] + "% → " + adjustFund[
            'afterTotalRatio'] + "%\t\n"
    strategy_ws.append(())
    strategy_ws.append(())
    strategy_ws.append(())
    strategy_ws.append(('【上次调仓动态】',))
    strategy_ws.append((detail,))

    strategy_ws.append(())
    strategy_ws.append(('【上次调仓详情】',))
    strategy_ws.append(('基金代码', '基金名称', '调整前占比', '调整后占比', '类型'))
    for adjustFund in adjust_list:
        for fund in adjustFund['fundList']:
            t = adjustFund['type']
            if t == 'a':
                t = '0'
            fund_type = fund_type_map[int(t)]
            strategy_ws.append((
                fund.get("fundCode", ''),
                fund.get("fundName", ''),
                fund['preRatio'] + "%",
                fund['afterRatio'] + "%",
                fund_type
            ))

    # 基金备选库，默认加载20个，剩余需再次发起ajax请求
    # strategyFundPool = content['strategyFundPool']['data']

    strategy_ws.append(())
    strategy_ws.append(())
    strategy_ws.append(())
    strategy_ws.append(('【成立以来日涨幅】',))
    strategy_ws.append(('日期', '组合涨幅', '基准涨幅'))
    # 成立来业绩表现
    total_performance = content['strategyProfitChart_ln'].get('data', {})
    for profit in total_performance:
        strategy_ws.append((
            profit['PDATE'] + '%',
            profit['SE'] + '%',
            profit['BENCH_SE'] + '%'
        ))
    strategy_name = content['baseData']['data']['strategyName']
    log.info('\t' + strategy_name + '详细信息爬取成功')
    strategy_ws.title = strategy_name


def parse_strategy_all(strategy_id, item, partner_ws, strategy_ws, content):
    is_success = parse_strategy_basic(strategy_id, item, partner_ws, content)
    if is_success:
        parse_strategy_detail(strategy_ws, content)


if __name__ == '__main__':
    root_dir = 'fundData/'
    skip_count = 0
    start = time.time()
    partnerIdList = get_partner_id_list()
    parse_count = 0
    holdWarehouseWb = None
    holdWarehouseWs = None
    wb = openpyxl.Workbook()
    ws = None
    mode = 'all'
    path_data = 'fundData\投顾策略详情'

    if len(sys.argv) >= 2:
        input_model = sys.argv[1].lower()
        if input_model == 'basic':
            mode = 'basic'
        if input_model == 'detail':
            log.info("正在清理上次爬取的数据文件...")
            del_file(path_data)
            mode = 'detail'
            wb = None
        if input_model == 'all':
            log.info("正在清理上次爬取的数据文件...")
            del_file(path_data)
    else:
        log.info("正在清理上次爬取的数据文件...")
        del_file(path_data)
    model_map = {
        'basic': '基本',
        'detail': '详细',
        'all': '全部'
    }

    log.info("【初始化完成，开始爬取" + model_map[mode] + "投顾策略信息】")
    for partnerId in partnerIdList:
        strategyIdSet = get_strategy_id_set(partnerId)
        baseInfo = get_base_info(partnerId)
        if wb is not None:
            ws = wb.create_sheet(baseInfo['partnerName'])
            ws.append((
                '公司名称',
                '投顾策略名称',
                '成立以来收益',
                '参考基准',
                '昨日涨幅',
                '成立天数',
                '建议持有',
                '年化投资目标',
                '上月涨幅',
                '参考基准',
                '近三月涨幅',
                '参考基准',
                '近半年涨幅',
                '参考基准',
                '近一年涨幅',
                '参考基准',
                '转入费率信息',
                '服务费率',
                '最近调仓时间',
                '调仓原因')
            )
        if mode != 'basic':
            holdWarehouseWb = openpyxl.Workbook()
        for strategyId in strategyIdSet:
            if mode != 'basic':
                holdWarehouseWs = holdWarehouseWb.create_sheet('Sheet')
            strategy_info = get_strategy_info(baseInfo, strategyId)
            if mode == 'basic':
                parse_strategy_basic(strategyId, baseInfo, ws, strategy_info)
            if mode == 'detail':
                parse_strategy_detail(holdWarehouseWs, strategy_info)
            if mode == 'all':
                parse_strategy_all(strategyId, baseInfo, ws, holdWarehouseWs, strategy_info)
        path = os.path.join(root_dir, '投顾策略详情')
        if not os.path.isdir(path):
            os.makedirs(path)
        if holdWarehouseWb is not None:
            holdWarehouseWs = holdWarehouseWb['Sheet']
            holdWarehouseWb.remove(holdWarehouseWs)
            holdWarehouseWb.save(path + '/' + baseInfo['partnerName'] + '产品详情.xlsx')
            holdWarehouseWb.close()
        parse_count += len(strategyIdSet)
    if not os.path.exists(root_dir):
        os.makedirs(root_dir)
    if wb is not None:
        ws = wb['Sheet']
        wb.remove(ws)
        timeStamp = get_timestamp()
        wb.save(root_dir + '投顾产品基本信息' + timeStamp + '.xlsx')
        wb.close()
    log.info("【爬取完毕】：成功爬取" + str(parse_count) + "条投顾策略")
    log.warning("其中" + str(skip_count) + "条投顾策略的基本数据为空，无法解析")
    log.info("共计爬取" + str(parse_count + skip_count) + "条投顾策略")
    log.info("耗时：" + str(round(time.time() - start, 3)) + "秒。")
